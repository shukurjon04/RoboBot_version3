
import csv
import io
import logging
import asyncio
import os
import openpyxl
from datetime import datetime
from typing import List
from aiogram import Router, F
from aiogram.filters import Command, CommandObject
from aiogram.types import Message, BufferedInputFile, CallbackQuery, FSInputFile
from aiogram.exceptions import TelegramRetryAfter, TelegramForbiddenError
from aiogram.fsm.context import FSMContext
from sqlalchemy import select, update
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.config.settings import settings
from app.infrastructure.repositories.sqlalchemy import SQLAlchemyUserRepository, SQLAlchemyChannelRepository
from app.infrastructure.database.models import WebinarSettings, User, Channel, WebinarCheckin
from app.utils.formatters import format_uzb_time
from app.presentation.keyboards.admin import admin_kb, admin_back_kb, suspicious_users_kb, checkin_button_kb
from app.presentation.keyboards.admin_channels import channels_list_kb, back_to_channels_kb
from app.domain.enums import UserStatus
from app.presentation.states import AdminSG
from app.infrastructure.telegram.checker import TelegramChannelChecker
from app.use_cases.subscription import SubscriptionService
from app.presentation.keyboards.registration import check_subscription_kb
from app.presentation.keyboards.admin_webinar import (
    webinar_years_kb, webinar_months_kb, webinar_days_kb, 
    webinar_hours_kb, webinar_minutes_kb
)

logger = logging.getLogger(__name__)

router = Router()

# Limit concurrent broadcast sends to avoid flood limits (20 msgs/sec is safer)
broadcast_semaphore = asyncio.Semaphore(20)

def is_admin(user_id: int) -> bool:
    return user_id in settings.ADMIN_IDS

@router.message(Command("admin"))
async def admin_panel(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Siz admin emassiz!")
        return
    
    text = (
        "👨‍💻 <b>Admin Panel</b>\n\n"
        "Quyidagi bo'limlardan birini tanlang:"
    )
    await message.answer(text, parse_mode="HTML", reply_markup=admin_kb)

@router.message(F.text == "⬅️ Orqaga")
async def admin_back_to_main(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    
    await state.clear()
    text = (
        "👨‍💻 <b>Admin Panel</b>\n\n"
        "Quyidagi bo'limlardan birini tanlang:"
    )
    await message.answer(text, parse_mode="HTML", reply_markup=admin_kb)

@router.callback_query(F.data == "back_to_admin_main")
async def on_back_to_admin_main(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    
    await state.clear()
    text = (
        "👨‍💻 <b>Admin Panel</b>\n\n"
        "Quyidagi bo'limlardan birini tanlang:"
    )
    await callback.message.delete()
    await callback.message.answer(text, parse_mode="HTML", reply_markup=admin_kb)

@router.message(F.text == "🏠 Asosiy menyu")
async def back_to_main(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    
    await state.clear()
    from app.presentation.keyboards.main import main_menu_kb
    await message.answer("Asosiy menyu:", reply_markup=main_menu_kb())

@router.message(F.text == "📢 Rassilka")
async def broadcast_button(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    
    await state.set_state(AdminSG.wait_broadcast)
    await message.answer(
        "✍️ <b>Rassilka xabarini yuboring</b>\n\n"
        "Xabar matn, rasm, video yoki boshqa fayl ko'rinishida bo'lishi mumkin. "
        "Yuborgan xabaringiz barcha foydalanuvchilarga aynan qanday bo'lsa shunday yetib boradi.",
        parse_mode="HTML",
        reply_markup=admin_back_kb()
    )

@router.message(AdminSG.wait_broadcast)
async def process_broadcast(message: Message, state: FSMContext, session):
    if not is_admin(message.from_user.id):
        return
    
    if message.text == "⬅️ Orqaga":
        await admin_back_to_main(message, state)
        return

    user_repo = SQLAlchemyUserRepository(session)
    users = await user_repo.get_all_users()
    
    await message.answer(
        "📤 <b>Rassilka boshlandi!</b>\n\n"
        f"Xabar {len(users)} ta foydalanuvchiga yuboriladi. "
        "Jarayon yakunlangach sizga hisobot beraman.",
        parse_mode="HTML"
    )
    
    # Run broadcast in background
    asyncio.create_task(_run_manual_broadcast(message, users, message.from_user.id))
    
    await state.clear()
    await message.answer("Boshqa amallar uchun menyudan foydalanishingiz mumkin:", reply_markup=admin_kb)

async def _run_manual_broadcast(message_to_copy: Message, users: List[User], admin_id: int):
    """Background task to perform parallel broadcast"""
    sent_count = 0
    blocked_count = 0
    error_count = 0
    admin_ids = set(settings.ADMIN_IDS)
    
    async def send_to_user(user: User):
        nonlocal sent_count, blocked_count, error_count
        if user.telegram_id in admin_ids:
            return

        async with broadcast_semaphore:
            try:
                # We use copy_to to preserve media and formatting
                await message_to_copy.copy_to(chat_id=user.telegram_id)
                sent_count += 1
            except TelegramForbiddenError:
                blocked_count += 1
            except TelegramRetryAfter as e:
                await asyncio.sleep(e.retry_after)
                try:
                    await message_to_copy.copy_to(chat_id=user.telegram_id)
                    sent_count += 1
                except Exception:
                    error_count += 1
            except Exception:
                error_count += 1

    logger.info(f"Admin {admin_id} started manual broadcast to {len(users)} users")
    tasks = [send_to_user(user) for user in users]
    await asyncio.gather(*tasks)
    
    logger.info(f"Manual broadcast finished. Success: {sent_count}, Blocked: {blocked_count}, Errors: {error_count}")
    
    try:
        report = (
            "✅ <b>Rassilka yakunlandi!</b>\n\n"
            f"👤 Jami: {len(users)}\n"
            f"✅ Yuborildi: {sent_count}\n"
            f"🚫 Bloklagan: {blocked_count}\n"
            f"❌ Xatoliklar: {error_count}"
        )
        await message_to_copy.bot.send_message(admin_id, report, parse_mode="HTML")
    except Exception as e:
        logger.error(f"Failed to send broadcast report to admin {admin_id}: {e}")

@router.message(F.text == "📊 Reyting Excel")
async def export_excel(message: Message, session):
    if not is_admin(message.from_user.id):
        return
        
    user_repo = SQLAlchemyUserRepository(session)
    users = await user_repo.get_top_users_by_balance(limit=10000)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reyting"
    
    # Headers with style
    headers = ["Rank", "ID", "Ism", "Username", "Ballar", "Viloyat", "Telefon"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for idx, u in enumerate(users, 1):
        ws.append([
            idx,
            u.id,
            u.full_name or u.first_name,
            u.username or "N/A",
            u.balance,
            u.region or "N/A",
            u.phone_number or "N/A"
        ])
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    document = BufferedInputFile(output.getvalue(), filename="reyting.xlsx")
    await message.answer_document(document, caption="📊 <b>Reyting (Excel)</b>", parse_mode="HTML")

@router.message(F.text == "⚠️ Shubhali foydalanuvchilar")
async def suspicious_users(message: Message, session):
    if not is_admin(message.from_user.id):
        return
    
    stmt = select(User).where(User.full_name == None).order_by(User.created_at.desc()).limit(20)
    result = await session.execute(stmt)
    suspicious = result.scalars().all()
    
    if not suspicious:
        await message.answer("✅ Shubhali foydalanuvchilar topilmadi.")
        return
    
    text = "⚠️ <b>Shubhali foydalanuvchilar:</b>\n\n"
    for u in suspicious:
        text += f"ID: {u.id} | TG: {u.telegram_id} | {u.first_name} | Ball: {u.balance}\n"
    
    text += "\n💡 Ularni bloklash uchun: /block [telegram_id]"
    text += "\n💡 Ballarni 0 ga: /reset [telegram_id]"
    text += "\n💡 Xabar yuborish uchun: /send [telegram_id]"
    
    await message.answer(text, parse_mode="HTML", reply_markup=suspicious_users_kb())

@router.callback_query(F.data == "send_to_suspicious")
async def ask_suspicious_broadcast(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
        
    await state.set_state(AdminSG.wait_suspicious_broadcast_content)
    await callback.message.answer(
        "✍️ <b>Barcha shubhali foydalanuvchilarga xabar yuborish</b>\n\n"
        "Xabar matn, rasm, video yoki boshqa fayl ko'rinishida bo'lishi mumkin.\n"
        "Yuborishni bekor qilish uchun /admin buyrug'ini bosing.",
        parse_mode="HTML",
        reply_markup=admin_back_kb()
    )
    await callback.answer()

@router.message(AdminSG.wait_suspicious_broadcast_content)
async def process_suspicious_broadcast(message: Message, state: FSMContext, session):
    if not is_admin(message.from_user.id):
        return
        
    if message.text == "⬅️ Orqaga":
        await admin_back_to_main(message, state)
        return

    # Fetch all suspicious users
    stmt = select(User).where(User.full_name == None)
    result = await session.execute(stmt)
    users = result.scalars().all()
    
    if not users:
        await message.answer("❌ Shubhali foydalanuvchilar topilmadi.")
        await state.clear()
        return

    await message.answer(
        f"📤 <b>Rassilka boshlandi!</b>\n"
        f"Xabar {len(users)} ta shubhali foydalanuvchiga yuboriladi.",
        parse_mode="HTML"
    )
    
    # Reuse the manual broadcast logic function
    asyncio.create_task(_run_manual_broadcast(message, users, message.from_user.id))
    
    await state.clear()
    await message.answer("Jarayon fonda davom etmoqda...", reply_markup=admin_kb)

@router.message(Command("block"))
async def block_user(message: Message, command: CommandObject, session):
    if not is_admin(message.from_user.id):
        return
    
    if not command.args:
        await message.answer("Foydalanuvchi ID sini kiriting: /block 123456789")
        return
    
    try:
        telegram_id = int(command.args)
        stmt = update(User).where(User.telegram_id == telegram_id).values(status=UserStatus.BLOCKED)
        await session.execute(stmt)
        await session.commit()
        await message.answer(f"✅ Foydalanuvchi {telegram_id} bloklandi.")
    except Exception as e:
        await message.answer(f"❌ Xato: {e}")

@router.message(Command("reset"))
async def reset_balance(message: Message, command: CommandObject, session):
    if not is_admin(message.from_user.id):
        return
    
    if not command.args:
        await message.answer("Foydalanuvchi ID sini kiriting: /reset 123456789")
        return
    
    try:
        telegram_id = int(command.args)
        stmt = update(User).where(User.telegram_id == telegram_id).values(balance=0)
        await session.execute(stmt)
        await session.commit()
        await message.answer(f"✅ Foydalanuvchi {telegram_id} ballari 0 ga tushirildi.")

    except Exception as e:
        await message.answer(f"❌ Xato: {e}")

@router.message(F.text == "✅ Check-in")
async def checkin_ask_text(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
        
    await state.set_state(AdminSG.wait_checkin_text)
    await message.answer(
        "✍️ <b>Check-in xabar matnini kiriting:</b>\n"
        "<i>Default: Efirda qatnashayotgan bo'lsangiz, tugmani bosing</i>\n\n"
        "Yoki 'skip' deb yozing (standart matn uchun).",
        parse_mode="HTML",
        reply_markup=admin_back_kb()
    )

@router.message(AdminSG.wait_checkin_text)
async def process_checkin_text(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    if message.text == "⬅️ Orqaga":
        await admin_back_to_main(message, state)
        return
        
    text = message.text
    if text.lower() == "skip":
        text = "Efirda qatnashayotgan bo'lsangiz, tugmani bosing 👇"
    
    await state.update_data(checkin_text=text)
    await state.set_state(AdminSG.wait_checkin_channel)
    
    await message.answer(
        "🔗 <b>Qaysi kanalga yuborilsin?</b>\n\n"
        "Kanalning <b>Username</b>ini (masalan: @kanalim), <b>ID</b>sini, yoki <b>Link</b>ini yuboring.\n"
        "<i>Bot o'sha kanalda admin bo'lishi kerak!</i>",
        parse_mode="HTML"
    )

@router.message(AdminSG.wait_checkin_channel)
async def process_checkin_channel(message: Message, state: FSMContext, bot):
    if not is_admin(message.from_user.id):
        return
        
    target = message.text.strip()
    
    # Try to extract username from link if present
    # Process link to extract username
    if "t.me/" in target:
        # Remove query parameters
        target = target.split('?')[0]
        # Remove trailing slash
        target = target.rstrip('/')
        
        parts = target.split('/')
        if parts:
            possible_username = parts[-1]
            if not possible_username.startswith('+') and not possible_username.isdigit(): 
                target = f"@{possible_username}"

    data = await state.get_data()
    text = data.get('checkin_text')
    
    bot_info = await bot.get_me()
    checkin_kb = checkin_button_kb(bot_info.username)
    
    try:
        await bot.send_message(chat_id=target, text=text, reply_markup=checkin_kb)
        await message.answer(f"✅ <b>Xabar muvaffaqiyatli yuborildi!</b>\nTarget: {target}", parse_mode="HTML", reply_markup=admin_kb)
    except Exception as e:
        await message.answer(f"❌ <b>Xatolik yuz berdi:</b>\n{e}\n\nBot kanalda admin ekanligini tekshiring.", parse_mode="HTML", reply_markup=admin_kb)
        
    await state.clear()

@router.message(Command("send"))
async def send_message_command(message: Message, command: CommandObject, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    
    # If ID provided in command
    if command.args:
        try:
            telegram_id = int(command.args)
            await state.update_data(send_message_id=telegram_id)
            await state.set_state(AdminSG.wait_send_message_content)
            await message.answer(
                f"✍️ <b>Foydalanuvchi ({telegram_id}) ga xabar yuboring:</b>\n\n"
                "Istalgan turdagi xabarni (matn, rasm, video, ovozli xabar) yuborishingiz mumkin.",
                parse_mode="HTML",
                reply_markup=admin_back_kb()
            )
        except ValueError:
            await message.answer("❌ ID raqam bo'lishi kerak!")
    else:
        # Ask for ID
        await state.set_state(AdminSG.wait_send_message_id)
        await message.answer(
            "✍️ <b>Xabar yuborish uchun foydalanuvchi ID sini kiriting:</b>",
            parse_mode="HTML",
            reply_markup=admin_back_kb()
        )

@router.message(AdminSG.wait_send_message_id)
async def process_send_message_id(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    if message.text == "⬅️ Orqaga":
        await admin_back_to_main(message, state)
        return

    try:
        telegram_id = int(message.text)
        await state.update_data(send_message_id=telegram_id)
        await state.set_state(AdminSG.wait_send_message_content)
        await message.answer(
            f"✍️ <b>Foydalanuvchi ({telegram_id}) ga xabar yuboring:</b>\n\n"
            "Istalgan turdagi xabarni (matn, rasm, video, ovozli xabar) yuborishingiz mumkin.",
            parse_mode="HTML",
            reply_markup=admin_back_kb()
        )
    except ValueError:
        await message.answer("❌ ID raqam bo'lishi kerak! Qaytadan kiriting:")

@router.message(AdminSG.wait_send_message_content)
async def process_send_message_content(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    if message.text == "⬅️ Orqaga":
        await admin_back_to_main(message, state)
        return

    data = await state.get_data()
    target_id = data.get('send_message_id')
    
    if not target_id:
        await message.answer("❌ Xatolik yuz berdi. Iltimos jarayonni boshidan boshlang.")
        await state.clear()
        return

    try:
        # Validate target ID type just in case
        target_id = int(target_id)
        
        # Copy the message content to the target user
        await message.copy_to(chat_id=target_id)
        
        await message.answer(
            f"✅ <b>Xabar yuborildi!</b>\n"
            f"ID: {target_id}", 
            parse_mode="HTML",
            reply_markup=admin_kb
        )
        await state.clear()
        
    except TelegramForbiddenError:
        await message.answer(f"🚫 Foydalanuvchi ({target_id}) botni bloklagan.")
        await state.clear()
    except Exception as e:
        await message.answer(f"❌ Yuborishda xatolik: {e}")
        await state.clear()

@router.message(F.text == "⏰ Vebinar vaqti")
async def set_webinar_time_button(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    
    await state.set_state(AdminSG.wait_webinar_year)
    await message.answer(
        "📅 <b>Vebinar yilini tanlang:</b>",
        parse_mode="HTML",
        reply_markup=webinar_years_kb()
    )

@router.callback_query(AdminSG.wait_webinar_year, F.data.startswith("wb_year:"))
async def process_webinar_year_cb(callback: CallbackQuery, state: FSMContext):
    year = callback.data.split(":")[1]
    await state.update_data(wb_year=int(year))
    await state.set_state(AdminSG.wait_webinar_month)
    await callback.message.edit_text(
        "📅 <b>Vebinar oyini tanlang:</b>",
        parse_mode="HTML",
        reply_markup=webinar_months_kb()
    )

@router.callback_query(AdminSG.wait_webinar_month, F.data.startswith("wb_month:"))
async def process_webinar_month_cb(callback: CallbackQuery, state: FSMContext):
    month = callback.data.split(":")[1]
    await state.update_data(wb_month=int(month))
    data = await state.get_data()
    year = data['wb_year']
    
    await state.set_state(AdminSG.wait_webinar_day)
    await callback.message.edit_text(
        "📅 <b>Vebinar kunini tanlang:</b>",
        parse_mode="HTML",
        reply_markup=webinar_days_kb(year, int(month))
    )

@router.callback_query(AdminSG.wait_webinar_day, F.data.startswith("wb_day:"))
async def process_webinar_day_cb(callback: CallbackQuery, state: FSMContext):
    day = callback.data.split(":")[1]
    await state.update_data(wb_day=int(day))
    
    await state.set_state(AdminSG.wait_webinar_hour)
    await callback.message.edit_text(
        "🕐 <b>Vebinar soatini tanlang:</b>",
        parse_mode="HTML",
        reply_markup=webinar_hours_kb()
    )

@router.callback_query(AdminSG.wait_webinar_hour, F.data.startswith("wb_hour:"))
async def process_webinar_hour_cb(callback: CallbackQuery, state: FSMContext):
    hour = callback.data.split(":")[1]
    await state.update_data(wb_hour=int(hour))
    
    await state.set_state(AdminSG.wait_webinar_minute)
    await callback.message.edit_text(
        "🕐 <b>Vebinar daqiqasini tanlang:</b>",
        parse_mode="HTML",
        reply_markup=webinar_minutes_kb(hour)
    )

@router.callback_query(AdminSG.wait_webinar_minute, F.data.startswith("wb_minute:"))
async def process_webinar_minute_cb(callback: CallbackQuery, state: FSMContext):
    minute = callback.data.split(":")[1]
    await state.update_data(wb_minute=int(minute))
    
    data = await state.get_data()
    year = data['wb_year']
    month = data['wb_month']
    day = data['wb_day']
    hour = data['wb_hour']
    
    webinar_dt = datetime(year, month, day, hour, int(minute))
    await state.update_data(webinar_dt=webinar_dt.isoformat())
    
    await state.set_state(AdminSG.wait_webinar_link)
    await callback.message.delete()
    await callback.message.answer(
        f"✅ <b>Vaqt belgilandi:</b> {format_uzb_time(webinar_dt)}\n\n"
        "Endi vebinar o'tkaziladigan kanal yoki guruh nomini (yoki havolasini) yuboring:",
        parse_mode="HTML",
        reply_markup=admin_back_kb()
    )

@router.callback_query(F.data == "wb_back_to_admin")
async def wb_back_to_admin_cb(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.delete()
    text = (
        "👨‍💻 <b>Admin Panel</b>\n\n"
        "Quyidagi bo'limlardan birini tanlang:"
    )
    await callback.message.answer(text, parse_mode="HTML", reply_markup=admin_kb)

# Callback Back Handlers
@router.callback_query(F.data == "wb_back_to_year")
async def wb_back_to_year(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AdminSG.wait_webinar_year)
    await callback.message.edit_text("📅 <b>Vebinar yilini tanlang:</b>", parse_mode="HTML", reply_markup=webinar_years_kb())

@router.callback_query(F.data == "wb_back_to_month")
async def wb_back_to_month(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AdminSG.wait_webinar_month)
    await callback.message.edit_text("📅 <b>Vebinar oyini tanlang:</b>", parse_mode="HTML", reply_markup=webinar_months_kb())

@router.callback_query(F.data == "wb_back_to_day")
async def wb_back_to_day(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    year = data['wb_year']
    month = data['wb_month']
    await state.set_state(AdminSG.wait_webinar_day)
    await callback.message.edit_text("📅 <b>Vebinar kunini tanlang:</b>", parse_mode="HTML", reply_markup=webinar_days_kb(year, month))

@router.callback_query(F.data == "wb_back_to_hour")
async def wb_back_to_hour(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AdminSG.wait_webinar_hour)
    await callback.message.edit_text("🕐 <b>Vebinar soatini tanlang:</b>", parse_mode="HTML", reply_markup=webinar_hours_kb())

@router.message(AdminSG.wait_webinar_link)
async def process_webinar_link(message: Message, state: FSMContext, session):
    if not is_admin(message.from_user.id):
        return
        
    if message.text == "⬅️ Orqaga":
        data = await state.get_data()
        hour = str(data.get('wb_hour', '00')).zfill(2)
        await state.set_state(AdminSG.wait_webinar_minute)
        await message.answer("🕐 <b>Vebinar daqiqasini tanlang:</b>", parse_mode="HTML", reply_markup=webinar_minutes_kb(hour))
        return

    data = await state.get_data()
    webinar_dt_iso = data.get('webinar_dt')
    if not webinar_dt_iso:
        await message.answer("Xatolik yuz berdi. Iltimos jarayonni qaytadan boshlang.")
        await state.clear()
        return
        
    webinar_dt = datetime.fromisoformat(webinar_dt_iso)
    webinar_link = message.text.strip()
    
    webinar = WebinarSettings(
        webinar_datetime=webinar_dt,
        webinar_link=webinar_link,
        sent_1h=False,
        sent_30m=False,
        sent_15m=False,
        sent_5m=False,
        sent_start=False
    )
    session.add(webinar)
    await session.commit()
    
    await state.clear()
    await message.answer(
        f"✅ <b>Vebinar muvaffaqiyatli rejalashtirildi!</b>\n\n"
        f"📅 {webinar_dt.strftime('%Y-%m-%d')}\n"
        f"🕐 {format_uzb_time(webinar_dt)}\n"
        f"🔗 Joylashuv: {webinar_link}\n\n"
        f"Eslatmalar quyidagi vaqtlarda yuboriladi:\n"
        f"• 1 soat oldin\n"
        f"• 30 daqiqa oldin\n"
        f"• 15 daqiqa oldin\n"
        f"• 5 daqiqa oldin\n"
        f"• Vebinar boshlanganda",
        parse_mode="HTML",
        reply_markup=admin_kb
    )

@router.message(F.text == "📢 Kanallarni boshqarish")
async def list_channels(message: Message, session):
    if not is_admin(message.from_user.id):
        return
    
    repo = SQLAlchemyChannelRepository(session)
    channels = await repo.get_all()
    
    await message.answer(
        "📢 <b>Kanallarni boshqarish</b>\n\n"
        "Quyida qo'shilgan kanallar ro'yxati keltirilgan. "
        "Kanalni o'chirish uchun uning nomini bosing:",
        parse_mode="HTML",
        reply_markup=channels_list_kb(channels)
    )

@router.callback_query(F.data == "back_to_channels_list")
async def on_back_to_channels_list(callback: CallbackQuery, session):
    if not is_admin(callback.from_user.id):
        return
    
    repo = SQLAlchemyChannelRepository(session)
    channels = await repo.get_all()
    
    await callback.message.edit_text(
        "📢 <b>Kanallarni boshqarish</b>\n\n"
        "Quyida qo'shilgan kanallar ro'yxati keltirilgan. "
        "Kanalni o'chirish uchun uning nomini bosing:",
        parse_mode="HTML",
        reply_markup=channels_list_kb(channels)
    )

@router.callback_query(F.data == "add_channel")
async def on_add_channel(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    
    await state.set_state(AdminSG.wait_channel_name)
    await callback.message.edit_text(
        "✍️ Kanal nomini kiriting (masalan: Robotronix):",
        reply_markup=back_to_channels_kb()
    )

@router.message(AdminSG.wait_channel_name)
async def process_channel_name(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    
    if message.text == "⬅️ Orqaga":
        await admin_back_to_main(message, state)
        return
        
    await state.update_data(channel_name=message.text)
    await state.set_state(AdminSG.wait_channel_id)
    await message.answer(
        "🆔 Kanal ID sini kiriting (masalan: -100123456789):\n"
        "<i>Bot ushbu kanalda admin bo'lishi shart!</i>",
        parse_mode="HTML",
        reply_markup=back_to_channels_kb()
    )

@router.message(AdminSG.wait_channel_id)
async def process_channel_id(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    
    if message.text == "⬅️ Orqaga":
        await admin_back_to_main(message, state)
        return
        
    await state.update_data(channel_id=message.text)
    await state.set_state(AdminSG.wait_channel_link)
    await message.answer(
        "🔗 Kanal havolasini kiriting (masalan: https://t.me/robotronix):",
        reply_markup=back_to_channels_kb()
    )

@router.message(AdminSG.wait_channel_link)
async def process_channel_link(message: Message, state: FSMContext, session, bot):
    if not is_admin(message.from_user.id):
        return
    
    if message.text == "⬅️ Orqaga":
        await admin_back_to_main(message, state)
        return
        
    data = await state.get_data()
    name = data.get("channel_name")
    ch_id = data.get("channel_id")
    link = message.text
    
    repo = SQLAlchemyChannelRepository(session)
    await repo.add_channel(ch_id, name, link)
    
    await state.clear()
    await message.answer(f"✅ Kanal '{name}' muvaffaqiyatli qo'shildi!", reply_markup=admin_kb)
    
    user_repo = SQLAlchemyUserRepository(session)
    users = await user_repo.get_all_users()
    
    checker = TelegramChannelChecker(bot)
    sub_service = SubscriptionService(repo, checker)
    
    broadcast_text = (
        f"📣 <b>Yangi kanal qo'shildi!</b>\n\n"
        f"Botdan foydalanishda davom etish uchun quyidagi kanallarga, jumladan <b>{name}</b> kanaliga obuna bo'lishingiz shart:"
    )
    
    count = 0
    admin_ids = set(settings.ADMIN_IDS)
    logger.info(f"Starting targeted broadcast for new channel '{name}'...")
    
    for u in users:
        # Strictly exclude admins
        if u.telegram_id in admin_ids:
            continue
            
        # Check if user needs to subscribe to anything
        is_subbed, unsubscribed = await sub_service.check_user_subscription(u.telegram_id)
        
        if not is_subbed:
            try:
                await bot.send_message(
                    u.telegram_id, 
                    broadcast_text, 
                    parse_mode="HTML", 
                    reply_markup=check_subscription_kb(unsubscribed)
                )
                count += 1
            except Exception:
                pass
            
    await message.answer(f"📢 Xabar faqat a'zo bo'lmagan {count} ta foydalanuvchiga yuborildi.")
    
    channels = await repo.get_all()
    await message.answer(
        "📢 Kanallar ro'yxati:",
        reply_markup=channels_list_kb(channels)
    )

@router.callback_query(F.data.startswith("del_channel:"))
async def on_delete_channel(callback: CallbackQuery, session):
    if not is_admin(callback.from_user.id):
        return
    
    channel_id = int(callback.data.split(":")[1])
    repo = SQLAlchemyChannelRepository(session)
    await repo.delete_channel(channel_id)
    
    await callback.answer("Kanal o'chirildi ✅")
    
    channels = await repo.get_all()
    await callback.message.edit_reply_markup(reply_markup=channels_list_kb(channels))

@router.message(F.text == "💾 Bazani yuklash")
async def backup_db(message: Message):
    if not is_admin(message.from_user.id):
        return
    
    from app.use_cases.backup import BackupService
    backup_service = BackupService()
    
    await message.answer("⏳ Baza yuklanmoqda, kuting...")
    
    try:
        file_content = await backup_service.create_backup()
        document = BufferedInputFile(file_content, filename=f"backup_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        await message.answer_document(document, caption="✅ Baza muvaffaqiyatli yuklandi!")
    except Exception as e:
        await message.answer(f"❌ Xatolik yuz berdi: {e}")

@router.message(F.text == "♻️ Bazani tiklash")
async def restore_db_ask(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    
    await state.set_state(AdminSG.wait_restore)
    await message.answer(
        "📂 <b>Bazani tiklash</b>\n\n"
        "Excel faylni (.xlsx) yuboring. \n"
        "⚠️ <b>DIQQAT:</b> Bu amal hozirgi bazadagi barcha ma'lumotlarni o'chirib yuboradi va yangi fayldagi ma'lumotlarni yozadi!",
        parse_mode="HTML",
        reply_markup=admin_back_kb()
    )

@router.message(AdminSG.wait_restore, F.document)
async def process_restore_db(message: Message, state: FSMContext, bot):
    if not is_admin(message.from_user.id):
        return
        
    document = message.document
    if not document.file_name.endswith('.xlsx'):
        await message.answer("❌ Faqat .xlsx formatidagi Excel faylni yuboring!")
        return
    
    await message.answer("⏳ Tiklash jarayoni boshlandi, kuting...")
    
    try:
        file_io = await bot.download(document)
        content = file_io.read()
        
        from app.use_cases.backup import BackupService
        backup_service = BackupService()
        await backup_service.restore_backup(content)
        
        await state.clear()
        await message.answer("✅ Baza muvaffaqiyatli tiklandi!", reply_markup=admin_kb)
    except Exception as e:
        await message.answer(f"❌ Xatolik yuz berdi: {e}")

@router.message(F.text == "📥 Vebinar qatnashchilari")
async def export_webinar_participants(message: Message, session):
    if not is_admin(message.from_user.id):
        return
        
    await message.answer("📥 Vebinar qatnashchilarini yuklab olinmoqda...")
    
    try:
        stmt = (
            select(WebinarCheckin, User)
            .join(User, WebinarCheckin.user_id == User.telegram_id)
            .order_by(WebinarCheckin.checked_at.desc())
        )
        result = await session.execute(stmt)
        records = result.all() # list of (WebinarCheckin, User)
        
        if not records:
            await message.answer("❌ Hali hech kim ro'yxatdan o'tmagan.")
            return

        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Qatnashchilar"
        
        # Header
        headers = ["ID", "Telegram ID", "Ism", "Username", "Telefon", "Hudud", "Yosh", "Status", "Check-in Vaqti"]
        ws.append(headers)
        
        for checkin, user in records:
            ws.append([
                user.id,
                user.telegram_id,
                user.full_name or user.first_name,
                f"@{user.username}" if user.username else "",
                user.phone_number,
                user.region,
                user.age_range,
                user.status,
                checkin.checked_at.strftime("%Y-%m-%d %H:%M:%S")
            ])
            
        filename = f"webinar_qatnashchilar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        file = FSInputFile(filename)
        await message.answer_document(file, caption="📊 Vebinar qatnashchilari ro'yxati")
        
        # Cleanup
        try:
            os.remove(filename)
        except:
            pass
    except Exception as e:
        logger.error(f"Export error: {e}")
        await message.answer(f"❌ Eksport xatoligi: {e}")

@router.message(F.text == "♻️ Vebinar tiklash")
async def ask_webinar_restore(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
        
    await state.set_state(AdminSG.wait_webinar_restore)
    await message.answer(
        "📂 <b>Vebinar qatnashchilarini tiklash</b>\n\n"
        "Excel faylni (.xlsx) yuboring.\n"
        "<i>Fayl quyidagi ustunlarga ega bo'lishi kerak: 'ID' (Telegram ID)</i>\n\n"
        "⚠️ <b>Eslatma:</b> Bu amal mavjud bazaga yangi qatnashchilarni qo'shadi. Takroriy ID lar o'tkazib yuboriladi.",
        parse_mode="HTML",
        reply_markup=admin_back_kb()
    )

@router.message(AdminSG.wait_webinar_restore, F.document)
async def process_webinar_restore(message: Message, state: FSMContext, bot, session):
    if not is_admin(message.from_user.id):
        return
        
    if message.text == "⬅️ Orqaga":
        await admin_back_to_main(message, state)
        return

    document = message.document
    if not document.file_name.endswith('.xlsx'):
        await message.answer("❌ Faqat .xlsx formatidagi Excel faylni yuboring!")
        return
    
    await message.answer("⏳ Tekshirish va tiklash jarayoni boshlandi, kuting...")
    
    try:
        file_io = await bot.download(document)
        wb = openpyxl.load_workbook(file_io)
        ws = wb.active
        
        added_count = 0
        skipped_count = 0
        
        # Headers are in row 1, data starts from row 2
        # We need to find the column index for "Telegram ID" or "ID"
        headers = [cell.value for cell in ws[1]]
        
        # Try to find telegram id column
        telegram_id_idx = -1
        for idx, h in enumerate(headers):
            if str(h).lower() in ["telegram id", "id", "user id", "userid"]:
                telegram_id_idx = idx
                break
                
        if telegram_id_idx == -1:
            await message.answer("❌ Excel faylda 'Telegram ID' yoki 'ID' ustuni topilmadi!")
            return
            
        # Iterate rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row: continue
            
            try:
                tg_id = row[telegram_id_idx]
                if not tg_id: continue
                
                tg_id = int(tg_id)
                
                # Check if exists
                stmt = select(WebinarCheckin).where(WebinarCheckin.user_id == tg_id)
                result = await session.execute(stmt)
                if result.scalars().first():
                    skipped_count += 1
                    continue
                    
                # Add checkin
                checkin = WebinarCheckin(user_id=tg_id, webinar_date=datetime.now())
                session.add(checkin)
                added_count += 1
                
            except Exception as row_err:
                logger.warning(f"Skipping row due to error: {row_err}")
                skipped_count += 1
                
        await session.commit()
        
        await state.clear()
        
        result_text = (
            "✅ <b>Tiklash yakunlandi!</b>\n\n"
            f"➕ Qo'shildi: {added_count}\n"
            f"⏭ O'tkazib yuborildi (mavjud): {skipped_count}"
        )
        await message.answer(result_text, parse_mode="HTML", reply_markup=admin_kb)
        
    except Exception as e:
        logger.error(f"Restore error: {e}", exc_info=True)
        await message.answer(f"❌ Xatolik yuz berdi: {e}")
